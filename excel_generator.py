# excel_generator.py
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation


# ----------------------------------------------------
# CATALOGUE (utilisé par app.py + Excel)
# ----------------------------------------------------
def get_catalog():
    panels = [
        ["Trina450", 450, 52.9, 44.6, 10.74, 10.09, -0.24],
        ["Trina505_DEG18", 505, 51.7, 43.7, 12.13, 11.56, -0.25],
        ["Solux415", 415, 37.95, 31.83, 13.77, 13.04, -0.28],
        ["Solux420", 420, 38.14, 32.02, 13.85, 13.12, -0.28],
        ["Solux425", 425, 38.32, 32.20, 13.93, 13.20, -0.28],
        ["TrinaS+475", 475, 39.0, 36.9, 14.72, 11.94, -0.24],
        ["TrinaS+480", 480, 39.2, 37.2, 14.77, 11.98, -0.24],
        ["TrinaS+485", 485, 39.4, 37.4, 14.84, 12.06, -0.24],
        ["TrinaS+490", 490, 39.6, 37.7, 14.91, 12.11, -0.24],
        ["TrinaS+495", 495, 39.8, 38.0, 14.97, 12.15, -0.24],
        ["TrinaS+500", 500, 40.1, 38.3, 15.03, 12.18, -0.24],
        ["TrinaS+505", 505, 40.3, 38.5, 15.09, 12.22, -0.24],
    ]

    inverters = [
        ("Sigen2.0", 2000, 4000, 50, 550, 600, 16, 2, "Mono"),
        ("Sigen3.0", 3000, 6000, 50, 550, 600, 16, 2, "Mono"),
        ("Sigen3.6", 3680, 7360, 50, 550, 600, 16, 2, "Mono"),
        ("Sigen4.0", 4000, 8000, 50, 550, 600, 16, 2, "Mono"),
        ("Sigen4.6", 4600, 9200, 50, 550, 600, 16, 2, "Mono"),
        ("Sigen5.0", 5000, 10000, 50, 550, 600, 16, 2, "Mono"),
        ("Sigen6.0", 6000, 12000, 50, 550, 600, 16, 2, "Mono"),

        ("Sigen3T", 3000, 6000, 160, 1000, 1100, 16, 2, "Tri 3x400"),
        ("Sigen4T", 4000, 8000, 160, 1000, 1100, 16, 2, "Tri 3x400"),
        ("Sigen5T", 5000, 10000, 160, 1000, 1100, 16, 2, "Tri 3x400"),
        ("Sigen6T", 6000, 12000, 160, 1000, 1100, 16, 2, "Tri 3x400"),
        ("Sigen8T", 8000, 16000, 160, 1000, 1100, 32, 2, "Tri 3x400"),
        ("Sigen10T", 10000, 20000, 160, 1000, 1100, 32, 2, "Tri 3x400"),
        ("Sigen12T", 12000, 24000, 160, 1000, 1100, 32, 2, "Tri 3x400"),
    ]

    batteries = [
        ["Sigen6", 6],
        ["Sigen10", 10],
    ]

    return panels, inverters, batteries


# ----------------------------------------------------
# FORMATTAGE AUTOMATIQUE
# ----------------------------------------------------
def _autofit(ws, width=16, max_col=25):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ----------------------------------------------------
# GENERATION EXCEL
# ----------------------------------------------------
def generate_workbook_bytes(config: dict) -> bytes:
    """
    Génère le fichier Excel complet et le renvoie en bytes pour Streamlit.
    """

    # Charger catalogue
    panels, inverters, batteries = get_catalog()

    wb = Workbook()

    # ------------------------------------------------
    # CATALOGUE
    # ------------------------------------------------
    ws = wb.active
    ws.title = "Catalogue"

    ws.append(["Panneaux"])
    ws.append(["ID", "P_STC_W", "Voc", "Vmp", "Isc", "Imp", "alpha_V_%/°C"])
    first_panel_row = ws.max_row + 1
    for p in panels:
        ws.append(p)
    last_panel_row = ws.max_row

    ws.append([""])
    ws.append(["Onduleurs"])
    ws.append(["ID", "P_AC_nom", "P_DC_max", "V_MPP_min", "V_MPP_max",
               "V_DC_max", "I_MPPT", "Nb_MPPT", "Type_reseau"])
    first_inv_row = ws.max_row + 1
    for inv in inverters:
        ws.append(list(inv))
    last_inv_row = ws.max_row

    ws.append([""])
    ws.append(["Batteries"])
    ws.append(["ID", "Cap_kWh"])
    first_bat_row = ws.max_row + 1
    for b in batteries:
        ws.append(b)
    last_bat_row = ws.max_row

    _autofit(ws, max_col=10)

    # ------------------------------------------------
    # CHOIX
    # ------------------------------------------------
    ws = wb.create_sheet("Choix")

    ws["A1"] = "Panneau"
    ws["A2"] = "Nombre modules"
    ws["A3"] = "Type réseau"
    ws["A4"] = "Batterie ?"
    ws["A5"] = "Batterie (kWh)"
    ws["A6"] = "Ratio DC/AC max"
    ws["A7"] = "Onduleur sélectionné"

    # Préremplir
    ws["B1"] = config.get("panel_id", "")
    ws["B2"] = config.get("n_modules", 10)
    ws["B3"] = config.get("grid_type", "")
    ws["B4"] = "Oui" if config.get("battery_enabled", False) else "Non"
    ws["B5"] = config.get("battery_kwh", 6)
    ws["B6"] = config.get("max_dc_ac", 1.3)
    ws["B7"] = config.get("inverter_id", "")

    # P_STC et puissance DC
    ws["A9"] = "P_STC panneau"
    ws["A10"] = "Puissance DC totale (W)"
    ws["B9"] = (
        f"=IFERROR(VLOOKUP(B1,Catalogue!$A${first_panel_row}:$G${last_panel_row},2,FALSE),\"\")"
    )
    ws["B10"] = "=IF(B9<>\"\",B9*B2,\"\")"

    _autofit(ws, max_col=7)

    # ------------------------------------------------
    # PROFIL
    # ------------------------------------------------
    ws = wb.create_sheet("Profil")

    ws["A1"] = "Conso annuelle (kWh)"
    ws["B1"] = config.get("annual_consumption", 3500)
    ws["A2"] = "Profil conso"
    ws["B2"] = config.get("consumption_profile", "Standard")

    ws.append([""])
    ws.append(["Mois", "%_conso", "Conso_kWh", "Prod_PV_kWh", "kWh_kWp_BEL", "Autocons_kWh"])

    months = ["Jan","Fév","Mar","Avr","Mai","Juin",
              "Juil","Août","Sep","Oct","Nov","Déc"]

    percent_std = [7,7,8,9,9,9,9,9,8,8,8,9]
    percent_winter = [10,10,10,9,8,7,6,6,7,8,9,10]
    percent_summer = [6,6,7,8,9,10,11,11,10,8,7,7]

    annual_kwh_kwp = 1034.0
    distribution = [3.8,5.1,8.7,11.5,12.1,11.8,11.9,10.8,9.7,7.0,4.3,3.3]
    kwh_kwp = [annual_kwh_kwp * d / 100 for d in distribution]

    start = 5
    for i, m in enumerate(months):
        r = start + i
        ws.cell(r, 1).value = m
        ws.cell(r, 2).value = (
            f"=CHOOSE(MATCH($B$2,{{\"Standard\",\"Hiver fort\",\"Été fort\"}},0),"
            f"{percent_std[i]},{percent_winter[i]},{percent_summer[i]})"
        )
        ws.cell(r, 3).value = f"=($B$1 * B{r} / 100)"
        ws.cell(r, 5).value = kwh_kwp[i]
        ws.cell(r, 4).value = f"=E{r} * Choix!$B$10 / 1000"
        ws.cell(r, 6).value = f"=MIN(C{r},D{r})"

    _autofit(ws, max_col=6)

    # Graph
    chart = BarChart()
    chart.title = "Production vs Consommation"
    data = Reference(ws, min_col=3, max_col=4, min_row=4, max_row=16)
    cats = Reference(ws, min_col=1, min_row=5, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    ws.add_chart(chart, "H4")

    # ------------------------------------------------
    # STRINGS
    # ------------------------------------------------
    ws = wb.create_sheet("Strings")

    ws["A1"] = "Vérification string"
    ws["A3"] = "Panneau"
    ws["A4"] = "Onduleur"
    ws["B3"] = config.get("panel_id", "")
    ws["B4"] = config.get("inverter_id", "")

    ws["A5"] = "T° min"
    ws["B5"] = config.get("t_min", -10)
    ws["A6"] = "T° max"
    ws["B6"] = config.get("t_max", 70)
    ws["A7"] = "Modules en série"
    ws["B7"] = config.get("n_series", 10)

    ws["A9"] = "Voc module"
    ws["A10"] = "Vmp module"
    ws["A11"] = "α_V (%/°C)"

    ws["B9"] = f"=VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},3,FALSE)"
    ws["B10"] = f"=VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},4,FALSE)"
    ws["B11"] = f"=VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},7,FALSE)"

    ws["A13"] = "V_DC_max"
    ws["A14"] = "V_MPP_min"
    ws["A15"] = "V_MPP_max"

    ws["B13"] = f"=VLOOKUP(B4,Catalogue!$A${first_inv_row}:$I${last_inv_row},6,FALSE)"
    ws["B14"] = f"=VLOOKUP(B4,Catalogue!$A${first_inv_row}:$I${last_inv_row},4,FALSE)"
    ws["B15"] = f"=VLOOKUP(B4,Catalogue!$A${first_inv_row}:$I${last_inv_row},5,FALSE)"

    ws["A17"] = "Voc string froid"
    ws["B17"] = "=B7 * B9 * (1 + B11/100*(B5-25))"

    ws["A18"] = "Vmp string chaud"
    ws["B18"] = "=B7 * B10 * (1 + B11/100*(B6-25))"

    ws["A20"] = "Check Voc <= V_DC_max"
    ws["B20"] = "=IF(B17<=B13,\"OK\",\"DÉPASSE\")"

    ws["A21"] = "Check Vmp dans MPPT"
    ws["B21"] = "=IF(AND(B18>=B14,B18<=B15),\"OK\",\"HORS PLAGE\")"

    _autofit(ws, max_col=4)

    # ------------------------------------------------
    # SYNTHESE
    # ------------------------------------------------
    ws = wb.create_sheet("Synthese")

    ws["A1"] = "Synthèse client"
    ws["A3"] = "Panneau"
    ws["B3"] = "=Choix!B1"

    ws["A4"] = "Modules"
    ws["B4"] = "=Choix!B2"

    ws["A5"] = "Puissance DC totale"
    ws["B5"] = "=Choix!B10"

    ws["A7"] = "Onduleur"
    ws["B7"] = "=Choix!B7"

    ws["A9"] = "Conso annuelle"
    ws["B9"] = "=Profil!B1"

    ws["A10"] = "Prod PV annuelle"
    ws["B10"] = "=SUM(Profil!D5:D16)"

    ws["A11"] = "Autocons annuelle"
    ws["B11"] = "=SUM(Profil!F5:F16)"

    ws["A12"] = "Taux autocons"
    ws["B12"] = "=IF(B10>0,B11/B10,\"\")"

    ws["A13"] = "Taux couverture"
    ws["B13"] = "=IF(B9>0,B11/B9,\"\")"

    ws["A15"] = "Batterie"
    ws["B15"] = "=Choix!B4"

    ws["A16"] = "Capacité batterie"
    ws["B16"] = "=Choix!B5"

    ws["A17"] = "Modèle batterie"
    ws["B17"] = "=Choix!B5"

    _autofit(ws, max_col=4)

    # ------------------------------------------------
    # SAUVEGARDE EN MÉMOIRE
    # ------------------------------------------------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
